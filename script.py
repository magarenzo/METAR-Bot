#!/usr/bin/python

import time
import datetime as dateTime
import xml.etree.ElementTree as elementTree
import urllib2
import openpyxl
from openpyxl import Workbook

# Use copy of original template
wb = openpyxl.load_workbook('./WeatherData.xlsx')
ws = wb.active

localtime = time.localtime(time.time())

# Download XML data
url = urllib2.urlopen('http://www.aviationweather.gov/adds/dataserver_current/httpparam?dataSource=metars&requestType=retrieve&format=xml&stationString=KJFK&mostRecent=true&hoursBeforeNow=1')
body = url.read()

# Extract and parse XML data; Add data to copied template
data = elementTree.fromstring(body)
d = data.findall('data')

for i in d:
        for j in i:
		hour = dateTime.datetime.now().hour
		hour += 2
		cell1 = 'A' + str(hour)
		cell2 = 'B' + str(hour)
		ws[cell1] = j.find('observation_time').text
		ws[cell2] = j.find('temp_c').text

# Save copied template
wb.save('./WeatherData.xlsx')
